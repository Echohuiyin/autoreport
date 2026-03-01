#!/usr/bin/env python3
"""
Script to check Excel file formatting, including background colors and font colors.
"""

import openpyxl

# Load the Excel file
excel_file = 'weekly_report.xlsx'
try:
    wb = openpyxl.load_workbook(excel_file, data_only=False)
    ws = wb.active
    print(f"Successfully loaded Excel file: {excel_file}")
    print(f"Sheet name: {ws.title}")
    print(f"Maximum row: {ws.max_row}")
    print(f"Maximum column: {ws.max_column}")
    
    # Check a few cells for formatting
    print("\nChecking cell formatting:")
    for row in range(1, 6):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            print(f"Cell {openpyxl.utils.get_column_letter(col)}{row}:")
            print(f"  Value: {cell.value}")
            try:
                print(f"  Fill: {cell.fill}")
                if cell.fill:
                    print(f"  Fill start color: {cell.fill.start_color}")
                    if hasattr(cell.fill.start_color, 'rgb') and cell.fill.start_color.rgb:
                        print(f"  Fill RGB: {cell.fill.start_color.rgb}")
            except Exception as e:
                print(f"  Error checking fill: {e}")
            try:
                print(f"  Font color: {cell.font.color}")
                if cell.font.color and hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
                    print(f"  Font RGB: {cell.font.color.rgb}")
            except Exception as e:
                print(f"  Error checking font color: {e}")
            print(f"  Font size: {cell.font.size}")
            print()
            
except Exception as e:
    print(f"Error: {e}")