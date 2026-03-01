#!/usr/bin/env python3
"""
Debug script to check directory structure and test Excel reading.
"""

import os
import sys

print("Python version:", sys.version)
print("Current directory:", os.getcwd())
print("Files in directory:")
for file in os.listdir('.'):
    print(f"  - {file}")
    print(f"    Size: {os.path.getsize(file)} bytes")
    print(f"    Exists: {os.path.exists(file)}")

# Try to import the modules
print("\nImporting modules...")
try:
    import pandas as pd
    import openpyxl
    print("Successfully imported pandas and openpyxl")
    print(f"pandas version: {pd.__version__}")
    print(f"openpyxl version: {openpyxl.__version__}")
except Exception as e:
    print(f"Error importing modules: {e}")

# Try to read the Excel file directly
print("\nTrying to read Excel file...")
excel_file = 'weekly_report.xlsx'
if os.path.exists(excel_file):
    print(f"Excel file exists: {excel_file}")
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
        print(f"Worksheet: {ws.title}")
        print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")
        
        # Read first few rows
        print("\nFirst 5 rows:")
        for i in range(1, min(6, ws.max_row + 1)):
            row = []
            for j in range(1, ws.max_column + 1):
                cell = ws.cell(row=i, column=j)
                row.append(cell.value)
            print(f"Row {i}: {row}")
        
        wb.close()
    except Exception as e:
        print(f"Error reading Excel file: {e}")
else:
    print(f"Excel file not found: {excel_file}")