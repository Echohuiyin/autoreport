#!/usr/bin/env python3
"""
Simple test script to verify Excel file can be read properly.
"""

import os
import pandas as pd
import openpyxl

print("Current directory:", os.getcwd())
print("Files in directory:", os.listdir('.'))

# Check if the Excel file exists
excel_file = 'weekly_report.xlsx'
if os.path.exists(excel_file):
    print(f"Excel file found: {excel_file}")
    
    # Try to read with pandas
    try:
        df = pd.read_excel(excel_file, header=1, engine='openpyxl')
        print(f"Successfully read Excel file with pandas")
        print(f"DataFrame shape: {df.shape}")
        print(f"Columns: {list(df.columns)}")
        print("\nFirst 5 rows:")
        print(df.head())
    except Exception as e:
        print(f"Error reading Excel file with pandas: {e}")
        
    # Try to read with openpyxl
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
        print(f"\nSuccessfully read Excel file with openpyxl")
        print(f"Worksheet title: {ws.title}")
        print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")
        wb.close()
    except Exception as e:
        print(f"Error reading Excel file with openpyxl: {e}")
else:
    print(f"Excel file not found: {excel_file}")