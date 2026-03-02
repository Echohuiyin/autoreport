#!/usr/bin/env python3
"""
Excel reader module for the weekly report system.
Handles reading Excel files and extracting data with formatting.
"""

import logging
import openpyxl
from src.exceptions import ExcelParsingError
from dataclasses import dataclass

logger = logging.getLogger(__name__)

@dataclass
class Cell:
    """Represents a cell with its value and formatting."""
    value: str
    rowspan: int = 1
    colspan: int = 1
    style: dict = None

@dataclass
class Row:
    """Represents a row of cells."""
    cells: list[Cell]

@dataclass
class ExcelData:
    """Represents the extracted Excel data."""
    title: str
    headers: list[str]
    rows: list[Row]

class ExcelReader:
    """Handles reading Excel files and extracting data with formatting."""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
    
    def read_excel_with_merged_cells(self):
        """
        Read Excel file and handle merged cells properly.
        This function fills in the missing values that result from merged cells.
        """
        try:
            # Load the workbook
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            ws = wb.active
            
            # Get merged cells
            merged_cells = ws.merged_cells.ranges
            
            # Create a dictionary to hold merged cell values
            merged_values = {}
            for merged_range in merged_cells:
                # Get the top-left cell value
                min_col, min_row, max_col, max_row = merged_range.bounds
                top_left_cell = ws.cell(row=min_row, column=min_col)
                value = top_left_cell.value
                
                # Store value for all cells in the merged range
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        merged_values[(row, col)] = value
            
            # Get the data
            data_rows = []
            headers = None
            
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row_idx == 1:
                    # Skip the title row "周报"
                    continue
                elif row_idx == 2:
                    # This is the header row
                    headers = [cell if cell is not None else f'Column_{i}' for i, cell in enumerate(row)]
                else:
                    # Data rows with merged cell handling
                    processed_row = []
                    for col_idx, cell_value in enumerate(row, 1):
                        # Check if this cell is part of a merged range
                        if (row_idx, col_idx) in merged_values:
                            processed_row.append(merged_values[(row_idx, col_idx)])
                        else:
                            processed_row.append(cell_value)
                    data_rows.append(processed_row)
            
            wb.close()
            
            if headers is None:
                raise ExcelParsingError("Could not find header row in Excel file")
            
            # Handle merged cells by forward-filling the '项目' column
            # This will propagate category names down through empty cells
            project_index = None
            for i, header in enumerate(headers):
                if header == '项目':
                    project_index = i
                    break
            
            if project_index is not None:
                current_project = None
                for row in data_rows:
                    if row[project_index] is not None:
                        current_project = row[project_index]
                    elif current_project is not None:
                        row[project_index] = current_project
            
            # Handle any remaining None values
            for row in data_rows:
                for i, value in enumerate(row):
                    if value is None:
                        row[i] = ''
            
            return headers, data_rows
            
        except Exception as e:
            logger.error(f"Error reading Excel file: {str(e)}")
            raise
    
    def read_excel_content(self) -> ExcelData:
        """
        Read and return the content of the Excel file as a structured data model
        with formatting preserved.
        """
        try:
            # Load the workbook with openpyxl to access formatting (data_only=False to preserve formatting)
            wb = openpyxl.load_workbook(self.file_path, data_only=False)
            ws = wb.active
            
            # Get merged cell ranges
            merged_cells = list(ws.merged_cells.ranges)
            
            # Create a dictionary to hold merged cell information
            merged_info = {}
            for merged_range in merged_cells:
                min_col, min_row, max_col, max_row = merged_range.bounds
                merged_info[(min_row, min_col)] = (max_row, max_col)
            
            # Read all columns from the Excel file
            column_indices = []
            headers = []
            
            # Check row 2 for headers
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=2, column=col)
                header_value = cell.value
                if header_value:
                    column_indices.append(col)
                    headers.append(header_value)
            
            # If no headers found, use default headers
            if not headers:
                column_indices = list(range(1, min(6, ws.max_column + 1)))
                headers = ['项目', '名称', '进展', '处理人', '状态'][:len(column_indices)]
            
            # Get title
            title_cell = ws.cell(row=1, column=1)
            title = title_cell.value if title_cell.value is not None else '周报'
            
            # Process data rows
            rows = []
            
            for row_idx in range(3, ws.max_row + 1):  # Start from row 3 (data rows)
                cells = []
                col_index = 0
                
                while col_index < len(column_indices):
                    col_idx = column_indices[col_index]
                    
                    # Check if this cell is part of a merged range that started in a previous row
                    skip_cell = False
                    for (start_row, start_col), (end_row, end_col) in merged_info.items():
                        if start_row < row_idx <= end_row and start_col <= col_idx <= end_col:
                            # This cell is part of a merged range from a previous row
                            skip_cell = True
                            break
                    
                    if skip_cell:
                        # Move to the next column
                        col_index += 1
                        continue
                    
                    # Check if this cell is the top-left of a new merged range
                    if (row_idx, col_idx) in merged_info:
                        max_row, max_col = merged_info[(row_idx, col_idx)]
                        rowspan = max_row - row_idx + 1
                        # Calculate colspan within our desired columns
                        colspan = 1
                        # Only count columns that are in our column_indices list
                        for i in range(col_idx + 1, max_col + 1):
                            if i in column_indices:
                                colspan += 1
                        # Ensure colspan doesn't exceed the number of columns we're displaying
                        colspan = min(colspan, len(column_indices) - col_index)
                    else:
                        rowspan = 1
                        colspan = 1
                    
                    # Get cell value
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value if cell.value is not None else ''
                    
                    # Get cell styling - preserve original formatting from Excel
                    style = {}
                    
                    # Background color
                    try:
                        # Check if cell has fill
                        if cell.fill and cell.fill.start_color:
                            # Get RGB value
                            fill_color = cell.fill.start_color
                            if fill_color and hasattr(fill_color, 'rgb'):
                                # Check if rgb is a string
                                if isinstance(fill_color.rgb, str):
                                    rgb_str = fill_color.rgb
                                    # Extract RGB part (remove alpha channel if present)
                                    if len(rgb_str) == 8:
                                        hex_color = '#' + rgb_str[2:]
                                    else:
                                        hex_color = '#' + rgb_str
                                    # Only set background color if it's not black (for readability)
                                    if hex_color != '#000000':
                                        style['background-color'] = hex_color
                    except Exception as e:
                        logger.debug(f"Error processing background color: {e}")
                    
                    # Font color
                    try:
                        # Check if cell has font color
                        if cell.font.color:
                            font_color = cell.font.color
                            if font_color:
                                # Handle theme-based colors
                                if hasattr(font_color, 'theme') and font_color.theme is not None:
                                    # For theme-based colors, use default black (theme 1 is usually black)
                                    if font_color.theme == 1:
                                        style['color'] = '#000000'
                                # Handle direct RGB colors
                                elif hasattr(font_color, 'rgb') and font_color.rgb:
                                    rgb_str = str(font_color.rgb)
                                    # Extract RGB part (remove alpha channel if present)
                                    if len(rgb_str) == 8:
                                        hex_color = '#' + rgb_str[2:]
                                    else:
                                        hex_color = '#' + rgb_str
                                    style['color'] = hex_color
                    except Exception as e:
                        logger.debug(f"Error processing font color: {e}")
                    
                    # Font size
                    if cell.font.size:
                        try:
                            style['font-size'] = f'{cell.font.size}pt'
                        except Exception as e:
                            logger.debug(f"Error processing font size: {e}")
                    
                    # Font weight
                    if cell.font.bold:
                        style['font-weight'] = 'bold'
                    
                    # Font style
                    if cell.font.italic:
                        style['font-style'] = 'italic'
                    
                    # Alignment
                    if cell.alignment:
                        if cell.alignment.horizontal:
                            style['text-align'] = cell.alignment.horizontal
                        if cell.alignment.vertical:
                            style['vertical-align'] = cell.alignment.vertical
                    
                    # Create cell object
                    cell_obj = Cell(value=str(value), rowspan=rowspan, colspan=colspan, style=style)
                    cells.append(cell_obj)
                    
                    # Move to the next column after the merged range
                    col_index += colspan
                
                # Create row object
                row_obj = Row(cells=cells)
                rows.append(row_obj)
            
            wb.close()
            
            # Create ExcelData object
            excel_data = ExcelData(title=title, headers=headers, rows=rows)
            
            # Log summary information
            logger.info(f"Successfully processed Excel file: {len(rows)} rows processed")
            logger.info(f"Found {len(merged_cells)} merged cell ranges")
            logger.info(f"Using {len(headers)} columns: {headers}")
            
            return excel_data
            
        except Exception as e:
            logger.error(f"Error reading Excel file: {str(e)}")
            raise