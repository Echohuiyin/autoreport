#!/usr/bin/env python3
"""
Test script for Excel reading functionality.
"""

import os
import sys
import random
import openpyxl

# Add the project root directory to the Python path
sys.path.insert(0, os.path.abspath(os.path.dirname(os.path.dirname(__file__))))

from src.excel.excel_reader import ExcelReader

print("Testing Excel reading functionality...")
print("=" * 60)

# Test 1: Basic Excel reading
try:
    print("\nTest 1: Basic Excel reading")
    print("-" * 40)
    
    # Create Excel reader instance
    reader = ExcelReader("weekly_report.xlsx")
    print("✓ ExcelReader instance created successfully")
    
    # Read Excel content
    html_content = reader.read_excel_content()
    print("✓ Excel content read successfully")
    print(f"  HTML length: {len(html_content)} characters")
    
    # Check if key columns are present
    if '项目' in html_content and '名称' in html_content and '进展' in html_content:
        print("✓ Key columns found in HTML output")
    else:
        print("✗ Key columns not found in HTML output")
    
except Exception as e:
    print(f"✗ Error: {e}")
    import traceback
    traceback.print_exc()

# Test 2: Randomly add M rows
try:
    print("\nTest 2: Randomly add M rows")
    print("-" * 40)
    
    # Create a copy of the original file
    wb = openpyxl.load_workbook('weekly_report.xlsx')
    ws = wb.active
    
    # Get current number of rows
    original_rows = ws.max_row
    print(f"  Original rows: {original_rows}")
    
    # Randomly add M rows (2-5 rows)
    M = random.randint(2, 5)
    print(f"  Adding {M} random rows")
    
    # Get header row
    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    
    # Add new rows
    for i in range(M):
        row_num = original_rows + i + 1
        ws.append(['测试项目', f'测试名称{i+1}', f'测试进展{i+1}', f'测试处理人{i+1}', f'测试备注{i+1}', f'测试状态{i+1}'])
    
    # Save the modified file
    test_file = 'test_with_extra_rows.xlsx'
    wb.save(test_file)
    print(f"  Added {M} rows to {test_file}")
    
    # Process the modified file
    reader = ExcelReader(test_file)
    html_content = reader.read_excel_content()
    print("✓ Modified file with extra rows processed successfully")
    
    # Clean up
    os.unlink(test_file)
    
except Exception as e:
    print(f"✗ Error: {e}")
    import traceback
    traceback.print_exc()

# Test 3: Randomly add N columns
try:
    print("\nTest 3: Randomly add N columns")
    print("-" * 40)
    
    # Create a copy of the original file
    wb = openpyxl.load_workbook('weekly_report.xlsx')
    ws = wb.active
    
    # Get current number of columns
    original_cols = ws.max_column
    print(f"  Original columns: {original_cols}")
    
    # Randomly add N columns (1-3 columns)
    N = random.randint(1, 3)
    print(f"  Adding {N} random columns")
    
    # Add new columns
    for i in range(N):
        col_num = original_cols + i + 1
        ws.insert_cols(col_num)
        ws.cell(row=1, column=col_num, value=f'测试列{i+1}')
        
        # Add data to the new column
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_num, value=f'测试数据{row-1}')
    
    # Save the modified file
    test_file = 'test_with_extra_columns.xlsx'
    wb.save(test_file)
    print(f"  Added {N} columns to {test_file}")
    
    # Process the modified file
    reader = ExcelReader(test_file)
    html_content = reader.read_excel_content()
    print("✓ Modified file with extra columns processed successfully")
    
    # Clean up
    os.unlink(test_file)
    
except Exception as e:
    print(f"✗ Error: {e}")
    import traceback
    traceback.print_exc()

# Test 4: Randomly add both M rows and N columns
try:
    print("\nTest 4: Randomly add both M rows and N columns")
    print("-" * 40)
    
    # Create a copy of the original file
    wb = openpyxl.load_workbook('weekly_report.xlsx')
    ws = wb.active
    
    # Get current dimensions
    original_rows = ws.max_row
    original_cols = ws.max_column
    print(f"  Original dimensions: {original_rows} rows × {original_cols} columns")
    
    # Randomly add N columns (1-2 columns)
    N = random.randint(1, 2)
    print(f"  Adding {N} random columns")
    
    # Add new columns
    for i in range(N):
        col_num = original_cols + i + 1
        ws.insert_cols(col_num)
        ws.cell(row=1, column=col_num, value=f'测试列{i+1}')
    
    # Randomly add M rows (2-3 rows)
    M = random.randint(2, 3)
    print(f"  Adding {M} random rows")
    
    # Add new rows
    for i in range(M):
        row_data = ['测试项目', f'测试名称{i+1}', f'测试进展{i+1}', f'测试处理人{i+1}', f'测试备注{i+1}', f'测试状态{i+1}']
        # Add data for new columns
        row_data.extend([f'测试数据{i+1}' for _ in range(N)])
        ws.append(row_data)
    
    # Save the modified file
    test_file = 'test_with_extra_rows_columns.xlsx'
    wb.save(test_file)
    print(f"  Added {M} rows and {N} columns to {test_file}")
    
    # Process the modified file
    reader = ExcelReader(test_file)
    html_content = reader.read_excel_content()
    print("✓ Modified file with extra rows and columns processed successfully")
    
    # Clean up
    os.unlink(test_file)
    
except Exception as e:
    print(f"✗ Error: {e}")
    import traceback
    traceback.print_exc()

print("\n" + "=" * 60)
print("Excel tests completed!")