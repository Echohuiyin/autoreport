#!/usr/bin/env python3
"""
Test script to verify the system can handle dynamic Excel files with randomly added rows or columns.
"""

import os
import random
import openpyxl
import subprocess
import sys

print("Testing dynamic Excel file handling...")

# Backup the original file
original_file = 'weekly_report.xlsx'
backup_file = 'weekly_report_backup.xlsx'

# Create backup
if os.path.exists(original_file):
    wb = openpyxl.load_workbook(original_file)
    wb.save(backup_file)
    print(f"Created backup: {backup_file}")
else:
    print(f"Error: {original_file} not found")
    sys.exit(1)

# Test 1: Add random columns
print("\nTest 1: Adding random columns")
try:
    # Load the Excel file
    wb = openpyxl.load_workbook(original_file)
    ws = wb.active
    
    # Add 2 new columns
    num_cols = 2
    last_col = ws.max_column
    
    for i in range(num_cols):
        # Insert new column
        ws.insert_cols(last_col + i + 1)
        # Add header
        new_header = f'新列{i+1}'
        ws.cell(row=2, column=last_col + i + 1, value=new_header)
        # Add data
        for row in range(3, ws.max_row + 1):
            ws.cell(row=row, column=last_col + i + 1, value=f'数据{row-2}')
    
    # Save changes
    wb.save(original_file)
    print(f"Added {num_cols} new columns")
    
    # Test the main program
    print("Running main program...")
    result = subprocess.run([sys.executable, 'weekly_report_sender.py'], capture_output=True, text=True)
    
    if result.returncode == 0:
        print("✓ Main program ran successfully with added columns")
    else:
        print("✗ Main program failed with added columns")
        print(f"Error: {result.stderr}")
        
    # Check if HTML was generated
    if os.path.exists('email_body.html'):
        with open('email_body.html', 'r', encoding='utf-8') as f:
            html_content = f.read()
        
        # Check if new columns are in HTML
        if '新列1' in html_content and '新列2' in html_content:
            print("✓ New columns are included in the HTML output")
        else:
            print("✗ New columns are not included in the HTML output")
    else:
        print("✗ No HTML output generated")
        
except Exception as e:
    print(f"Error: {e}")

# Restore original file
if os.path.exists(backup_file):
    wb = openpyxl.load_workbook(backup_file)
    wb.save(original_file)
    print(f"\nRestored original file")

# Test 2: Add random rows
print("\nTest 2: Adding random rows")
try:
    # Load the Excel file
    wb = openpyxl.load_workbook(original_file)
    ws = wb.active
    
    # Add 3 new rows
    num_rows = 3
    last_row = ws.max_row
    
    for i in range(num_rows):
        # Insert new row
        ws.insert_rows(last_row + i + 1)
        # Add data
        ws.cell(row=last_row + i + 1, column=1, value='测试项目')
        ws.cell(row=last_row + i + 1, column=2, value=f'测试名称{i+1}')
        ws.cell(row=last_row + i + 1, column=3, value='测试进展')
        ws.cell(row=last_row + i + 1, column=4, value='测试人员')
        ws.cell(row=last_row + i + 1, column=5, value='测试备注')
        ws.cell(row=last_row + i + 1, column=6, value='open')
    
    # Save changes
    wb.save(original_file)
    print(f"Added {num_rows} new rows")
    
    # Test the main program
    print("Running main program...")
    result = subprocess.run([sys.executable, 'weekly_report_sender.py'], capture_output=True, text=True)
    
    if result.returncode == 0:
        print("✓ Main program ran successfully with added rows")
    else:
        print("✗ Main program failed with added rows")
        print(f"Error: {result.stderr}")
        
    # Check if HTML was generated
    if os.path.exists('email_body.html'):
        with open('email_body.html', 'r', encoding='utf-8') as f:
            html_content = f.read()
        
        # Check if new rows are in HTML
        if '测试项目' in html_content and '测试名称1' in html_content:
            print("✓ New rows are included in the HTML output")
        else:
            print("✗ New rows are not included in the HTML output")
    else:
        print("✗ No HTML output generated")
        
except Exception as e:
    print(f"Error: {e}")

# Restore original file
if os.path.exists(backup_file):
    wb = openpyxl.load_workbook(backup_file)
    wb.save(original_file)
    print(f"\nRestored original file")

# Clean up backup
if os.path.exists(backup_file):
    os.remove(backup_file)
    print(f"Cleaned up backup file: {backup_file}")

print("\nAll tests completed!")