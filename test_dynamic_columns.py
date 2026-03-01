#!/usr/bin/env python3
"""
Test script to verify the system can handle dynamic Excel files with added columns.
"""

import os
import openpyxl
from weekly_report_sender import WeeklyReportSender

print("Testing dynamic Excel file handling...")

# Create a test file by copying the original
original_file = 'weekly_report.xlsx'
test_file = 'test_weekly_report.xlsx'

# Copy the original file
wb = openpyxl.load_workbook(original_file)
wb.save(test_file)
print(f"Created test file: {test_file}")

# Test 1: Original file
print("\nTest 1: Original file")
try:
    sender = WeeklyReportSender(excel_file_path=test_file)
    html_content = sender.read_excel_content()
    print("✓ Original file processed successfully")
    # Count columns
    header_start = html_content.find('<tr>')
    header_end = html_content.find('</tr>', header_start + 1)
    header_row = html_content[header_start:header_end]
    original_columns = header_row.count('<th>')
    print(f"  Original columns: {original_columns}")
except Exception as e:
    print(f"✗ Error: {e}")

# Test 2: Add columns
print("\nTest 2: Add 2 new columns")
try:
    # Load and modify the test file
    wb = openpyxl.load_workbook(test_file)
    ws = wb.active
    
    # Add 2 new columns
    last_col = ws.max_column
    for i in range(2):
        ws.insert_cols(last_col + i + 1)
        ws.cell(row=2, column=last_col + i + 1, value=f'新列{i+1}')
        for row in range(3, ws.max_row + 1):
            ws.cell(row=row, column=last_col + i + 1, value=f'数据{row-2}')
    
    wb.save(test_file)
    print("  Added 2 new columns")
    
    # Test processing
    sender = WeeklyReportSender(excel_file_path=test_file)
    html_content = sender.read_excel_content()
    print("✓ Modified file processed successfully")
    
    # Check if new columns are in HTML
    if '新列1' in html_content and '新列2' in html_content:
        print("✓ New columns are included in HTML")
    else:
        print("✗ New columns are not in HTML")
        
    # Count columns in modified file
    header_start = html_content.find('<tr>')
    header_end = html_content.find('</tr>', header_start + 1)
    header_row = html_content[header_start:header_end]
    modified_columns = header_row.count('<th>')
    print(f"  Modified columns: {modified_columns}")
    
    if modified_columns == original_columns + 2:
        print("✓ Column count is correct")
    else:
        print("✗ Column count is incorrect")
        
except Exception as e:
    print(f"✗ Error: {e}")

# Test 3: Add rows
print("\nTest 3: Add 3 new rows")
try:
    # Load and modify the test file
    wb = openpyxl.load_workbook(test_file)
    ws = wb.active
    
    # Add 3 new rows
    last_row = ws.max_row
    for i in range(3):
        ws.insert_rows(last_row + i + 1)
        ws.cell(row=last_row + i + 1, column=1, value='测试项目')
        ws.cell(row=last_row + i + 1, column=2, value=f'测试名称{i+1}')
        ws.cell(row=last_row + i + 1, column=3, value='测试进展')
        ws.cell(row=last_row + i + 1, column=4, value='测试人员')
        ws.cell(row=last_row + i + 1, column=5, value='测试备注')
        ws.cell(row=last_row + i + 1, column=6, value='open')
        ws.cell(row=last_row + i + 1, column=7, value=f'数据{last_row + i + 1}')
        ws.cell(row=last_row + i + 1, column=8, value=f'数据{last_row + i + 1}')
    
    wb.save(test_file)
    print("  Added 3 new rows")
    
    # Test processing
    sender = WeeklyReportSender(excel_file_path=test_file)
    html_content = sender.read_excel_content()
    print("✓ File with added rows processed successfully")
    
    # Check if new rows are in HTML
    if '测试项目' in html_content and '测试名称1' in html_content:
        print("✓ New rows are included in HTML")
    else:
        print("✗ New rows are not in HTML")
        
except Exception as e:
    print(f"✗ Error: {e}")

# Clean up
if os.path.exists(test_file):
    os.remove(test_file)
    print(f"\nCleaned up test file: {test_file}")

print("\nAll tests completed!")