#!/usr/bin/env python3
"""
Direct test for dynamic Excel columns handling.
"""

import os
import openpyxl
from weekly_report_sender import WeeklyReportSender

# Print current directory
print("Current directory:", os.getcwd())
# List files in current directory
print("Files in directory:", os.listdir('.'))

# Test 1: Check if original file exists
print("\nTest 1: Checking original file")
original_file = 'weekly_report.xlsx'
if os.path.exists(original_file):
    print(f"✓ Original file exists: {original_file}")
else:
    print(f"✗ Original file not found: {original_file}")

# Test 2: Create modified file with extra column
print("\nTest 2: Creating modified file with extra column")
try:
    # Load original file
    wb = openpyxl.load_workbook(original_file)
    ws = wb.active
    
    # Get current columns
    current_cols = ws.max_column
    print(f"  Current columns: {current_cols}")
    
    # Add a new column
    ws.insert_cols(current_cols + 1)
    ws.cell(row=2, column=current_cols + 1, value='测试列')
    
    # Add data to new column
    for row in range(3, ws.max_row + 1):
        ws.cell(row=row, column=current_cols + 1, value=f'测试数据{row-2}')
    
    # Save modified file
    test_file = 'test_modified.xlsx'
    wb.save(test_file)
    print(f"  Modified file saved: {test_file}")
    
    # Test 3: Process modified file
    print("\nTest 3: Processing modified file")
    sender = WeeklyReportSender(excel_file_path=test_file)
    html_content = sender.read_excel_content()
    print("✓ Modified file processed successfully")
    
    # Save HTML to file
    html_file = 'test_output.html'
    with open(html_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    print(f"  HTML output saved: {html_file}")
    
    # Check if new column is in HTML
    if '测试列' in html_content:
        print("✓ New column '测试列' is included in HTML")
    else:
        print("✗ New column '测试列' is not included in HTML")
    
    # Check if data is in HTML
    if '测试数据1' in html_content:
        print("✓ Data from new column is included in HTML")
    else:
        print("✗ Data from new column is not included in HTML")
    
    # Clean up
    if os.path.exists(test_file):
        os.remove(test_file)
        print(f"  Cleaned up: {test_file}")
        
except Exception as e:
    print(f"✗ Error: {e}")
    import traceback
    traceback.print_exc()

print("\nTest completed!")