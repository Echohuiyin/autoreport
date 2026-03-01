#!/usr/bin/env python3
"""
Test script to verify the system can handle dynamic Excel files with randomly added rows or columns.
"""

import os
import random
import openpyxl
from weekly_report_sender import WeeklyReportSender
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def modify_excel_file(original_file, output_file, add_rows=True, num_rows=3, num_cols=1):
    """
    Modify the Excel file by adding random rows or columns.
    
    Args:
        original_file: Path to the original Excel file
        output_file: Path to save the modified Excel file
        add_rows: Whether to add rows (True) or columns (False)
        num_rows: Number of rows to add
        num_cols: Number of columns to add
    """
    try:
        # Load the original workbook
        wb = openpyxl.load_workbook(original_file)
        ws = wb.active
        
        if add_rows:
            # Add rows
            logger.info(f"Adding {num_rows} random rows to Excel file")
            # Find the last row with data
            last_row = ws.max_row
            # Add new rows after the last data row
            for i in range(num_rows):
                # Insert a new row
                ws.insert_rows(last_row + i + 1)
                # Add some random data
                for col in range(1, ws.max_column + 1):
                    # Get the header of the column
                    header = ws.cell(row=2, column=col).value
                    # Add random data based on column header
                    if header == '项目':
                        ws.cell(row=last_row + i + 1, column=col, value='测试项目')
                    elif header == '名称':
                        ws.cell(row=last_row + i + 1, column=col, value=f'测试名称{i+1}')
                    elif header == '进展':
                        ws.cell(row=last_row + i + 1, column=col, value='测试进展')
                    elif header == '处理人':
                        ws.cell(row=last_row + i + 1, column=col, value='测试人员')
                    elif header == '备注':
                        ws.cell(row=last_row + i + 1, column=col, value=f'测试备注{i+1}')
                    elif header == '状态':
                        ws.cell(row=last_row + i + 1, column=col, value='open')
        else:
            # Add columns
            logger.info(f"Adding {num_cols} random columns to Excel file")
            # Find the last column with data
            last_col = ws.max_column
            # Add new columns after the last data column
            for i in range(num_cols):
                # Insert a new column
                ws.insert_cols(last_col + i + 1)
                # Add header
                new_header = f'新列{i+1}'
                ws.cell(row=2, column=last_col + i + 1, value=new_header)
                # Add random data to the new column
                for row in range(3, ws.max_row + 1):
                    ws.cell(row=row, column=last_col + i + 1, value=f'数据{row-2}')
        
        # Save the modified workbook
        wb.save(output_file)
        logger.info(f"Modified Excel file saved to {output_file}")
        
        return True
    except Exception as e:
        logger.error(f"Error modifying Excel file: {e}")
        return False

def test_dynamic_excel():
    """
    Test the system with dynamically modified Excel files.
    """
    original_file = 'weekly_report.xlsx'
    
    # Test 1: Add random rows
    logger.info("=== Test 1: Adding random rows ===")
    modified_file_rows = 'weekly_report_modified_rows.xlsx'
    if modify_excel_file(original_file, modified_file_rows, add_rows=True, num_rows=random.randint(2, 5)):
        # Test the report sender with the modified file
        try:
            sender = WeeklyReportSender(excel_file_path=modified_file_rows)
            html_content = sender.read_excel_content()
            logger.info("Successfully generated HTML with added rows")
            # Check if the HTML contains the test data
            if '测试项目' in html_content and '测试名称' in html_content:
                logger.info("✓ Test 1 passed: Added rows are included in the HTML")
            else:
                logger.warning("✗ Test 1 failed: Added rows are not included in the HTML")
        except Exception as e:
            logger.error(f"Error processing modified file with rows: {e}")
    
    # Test 2: Add random columns
    logger.info("\n=== Test 2: Adding random columns ===")
    modified_file_cols = 'weekly_report_modified_cols.xlsx'
    if modify_excel_file(original_file, modified_file_cols, add_rows=False, num_cols=random.randint(1, 3)):
        # Test the report sender with the modified file
        try:
            sender = WeeklyReportSender(excel_file_path=modified_file_cols)
            html_content = sender.read_excel_content()
            logger.info("Successfully generated HTML with added columns")
            # Check if the HTML contains the new column headers
            if '新列' in html_content:
                logger.info("✓ Test 2 passed: Added columns are included in the HTML")
            else:
                logger.warning("✗ Test 2 failed: Added columns are not included in the HTML")
        except Exception as e:
            logger.error(f"Error processing modified file with columns: {e}")
    
    # Clean up
    try:
        if os.path.exists(modified_file_rows):
            os.remove(modified_file_rows)
            logger.info(f"Cleaned up: {modified_file_rows}")
        if os.path.exists(modified_file_cols):
            os.remove(modified_file_cols)
            logger.info(f"Cleaned up: {modified_file_cols}")
    except Exception as e:
        logger.error(f"Error cleaning up: {e}")

if __name__ == "__main__":
    test_dynamic_excel()