#!/usr/bin/env python3
"""
Test script to verify dynamic Excel handling by generating a modified file and checking the output.
"""

import os
import openpyxl
from weekly_report_sender import WeeklyReportSender

print("Testing dynamic Excel handling...")

# Create a modified Excel file with extra column
print("Creating modified Excel file with extra column...")

# Load original file
wb = openpyxl.load_workbook('weekly_report.xlsx')
ws = wb.active

# Add a new column
last_col = ws.max_column
ws.insert_cols(last_col + 1)
ws.cell(row=2, column=last_col + 1, value='测试列')

# Add data to the new column
for row in range(3, ws.max_row + 1):
    ws.cell(row=row, column=last_col + 1, value=f'测试数据{row-2}')

# Save modified file
test_file = 'test_modified.xlsx'
wb.save(test_file)
print(f"Modified file saved as: {test_file}")

# Process the modified file
print("Processing modified file...")
sender = WeeklyReportSender(excel_file_path=test_file)
html_content = sender.read_excel_content()

# Save HTML output
with open('email_body_dynamic.html', 'w', encoding='utf-8') as f:
    f.write(html_content)
print("HTML output saved as: email_body_dynamic.html")

# Check if new column is in HTML
if '测试列' in html_content:
    print("✓ New column '测试列' is included in the HTML output")
else:
    print("✗ New column '测试列' is not included in the HTML output")

# Check for specific data
if '测试数据1' in html_content:
    print("✓ Data from new column is included in the HTML output")
else:
    print("✗ Data from new column is not included in the HTML output")

# Clean up
if os.path.exists(test_file):
    os.remove(test_file)
    print(f"Cleaned up test file: {test_file}")

print("Test completed!")