#!/usr/bin/env python3
"""
Test script to verify Excel file reading functionality with merged cells support.
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os
import sys

def test_excel_with_merged_cells():
    """Test reading Excel file with proper merged cell handling."""
    # Check for both possible filenames (with and without typo)
    possible_filenames = ['weekly_report.xlsx', 'weekyly_report.xlsx', 'weekly report.xlsx', 'weekyly report.xlsx']
    excel_file = None
    
    for filename in possible_filenames:
        if os.path.exists(filename):
            excel_file = filename
            break
    
    if excel_file is None:
        print("ERROR: No Excel file found!")
        print(f"Searched for: {possible_filenames}")
        print(f"Files in current directory: {[f for f in os.listdir('.') if f.endswith('.xlsx')]}")
        return False
    
    print(f"Found Excel file: {excel_file}")
    
    try:
        # Load workbook with openpyxl to access merged cell information
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        print(f"Worksheet dimensions: {ws.dimensions}")
        print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")
        
        # Get merged cell ranges
        merged_ranges = ws.merged_cells.ranges
        print(f"Merged cell ranges found: {len(merged_ranges)}")
        for merged_range in merged_ranges:
            print(f"  - {merged_range}")
        
        # Read data row by row to understand the actual structure
        print("\nActual cell values (row by row):")
        for row_idx in range(1, min(ws.max_row + 1, 20)):  # Show first 20 rows
            row_values = []
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_values.append(str(cell_value) if cell_value is not None else 'None')
            print(f"Row {row_idx}: {row_values}")
        
        # Now try to read with pandas but handle the structure properly
        print("\n" + "="*50)
        print("Attempting to parse structured data...")
        
        # Skip the title row (row 1) and use row 2 as headers
        df = pd.read_excel(excel_file, header=1, engine='openpyxl')
        print(f"DataFrame shape: {df.shape}")
        print(f"Columns: {list(df.columns)}")
        
        # Show the DataFrame
        print("\nDataFrame content:")
        print(df.to_string())
        
        wb.close()
        return True
        
    except Exception as e:
        print(f"ERROR: Failed to read Excel file '{excel_file}': {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def test_simple_pandas():
    """Test simple pandas reading for comparison."""
    possible_filenames = ['weekly_report.xlsx', 'weekyly_report.xlsx', 'weekly report.xlsx', 'weekyly report.xlsx']
    excel_file = None
    
    for filename in possible_filenames:
        if os.path.exists(filename):
            excel_file = filename
            break
    
    if excel_file is None:
        return False
    
    try:
        df = pd.read_excel(excel_file, header=1)
        print("Simple pandas reading:")
        print(f"Shape: {df.shape}")
        print(df.head())
        return True
    except Exception as e:
        print(f"Simple pandas error: {e}")
        return False

if __name__ == "__main__":
    print("Testing Excel file with merged cells support...")
    print(f"Current directory: {os.getcwd()}")
    print(f"Files in directory: {os.listdir('.')}")
    success = test_excel_with_merged_cells()
    print(f"Test result: {'SUCCESS' if success else 'FAILED'}")
    sys.exit(0 if success else 1)