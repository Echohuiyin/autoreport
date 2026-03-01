#!/usr/bin/env python3
"""
Test script that writes output to a file to verify dynamic Excel handling.
"""

import os
import openpyxl
from weekly_report_sender import WeeklyReportSender

# Write to output file
with open('test_output.txt', 'w', encoding='utf-8') as f:
    f.write("Testing dynamic Excel handling...\n\n")
    
    # Test 1: Original file
    f.write("Test 1: Original file\n")
    try:
        sender = WeeklyReportSender()
        html_content = sender.read_excel_content()
        f.write("✓ Original file processed successfully\n")
        
        # Check for existing columns
        if '项目' in html_content and '名称' in html_content and '进展' in html_content:
            f.write("✓ All original columns found\n")
        else:
            f.write("✗ Some original columns missing\n")
    except Exception as e:
        f.write(f"✗ Error: {e}\n")
    
    f.write("\nTest 2: Modified file with extra column\n")
    try:
        # Import config to modify it temporarily
        from config import FILE_CONFIG
        
        # Create a copy of the original file
        wb = openpyxl.load_workbook('weekly_report.xlsx')
        ws = wb.active
        
        # Add a new column
        last_col = ws.max_column
        ws.insert_cols(last_col + 1)
        ws.cell(row=2, column=last_col + 1, value='测试列')
        
        # Add data to the new column
        for row in range(3, ws.max_row + 1):
            ws.cell(row=row, column=last_col + 1, value=f'测试数据{row-2}')
        
        # Save the modified file
        test_file = 'test_with_extra_column.xlsx'
        wb.save(test_file)
        f.write(f"  Added extra column '测试列'\n")
        
        # Temporarily update the FILE_CONFIG to use our test file
        original_file_path = FILE_CONFIG['excel_file_path']
        FILE_CONFIG['excel_file_path'] = test_file
        
        # Process the modified file
        sender = WeeklyReportSender()
        # Skip validation since we know the file exists
        # sender.validate_config()
        # Manually set the file path
        sender.excel_file_path = test_file
        html_content = sender.read_excel_content()
        f.write("✓ Modified file processed successfully\n")
        
        # Check if the new column is included
        if '测试列' in html_content:
            f.write("✓ New column '测试列' is included in the output\n")
        else:
            f.write("✗ New column '测试列' is not included in the output\n")
            
        # Check for data
        if '测试数据1' in html_content:
            f.write("✓ Data from new column is included\n")
        else:
            f.write("✗ Data from new column is not included\n")
        
        # Restore original file path
        FILE_CONFIG['excel_file_path'] = original_file_path
        
        # Clean up
        if os.path.exists(test_file):
            os.remove(test_file)
            f.write("  Cleaned up test file\n")
    except Exception as e:
        f.write(f"✗ Error: {e}\n")
        import traceback
        f.write(traceback.format_exc())
    
    f.write("\nTest completed!\n")

print("Test output written to test_output.txt")