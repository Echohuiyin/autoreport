#!/usr/bin/env python3
"""
Test script to verify the system can handle dynamic Excel files with randomly added rows or columns.
"""

import os
import random
import openpyxl
from weekly_report_sender import WeeklyReportSender

print("Testing random Excel file changes...")

# Create a test file by copying the original
original_file = 'weekly_report.xlsx'
test_file = 'test_weekly_report.xlsx'

# Copy the original file
wb = openpyxl.load_workbook(original_file)
wb.save(test_file)
print(f"Created test file: {test_file}")

# Test 1: Original file
print("\nTest 1: Original file")
try:
    sender = WeeklyReportSender(excel_file_path=test_file)
    html_content = sender.read_excel_content()
    print("✓ Original file processed successfully")
    # Count columns
    header_start = html_content.find('<tr>')
    header_end = html_content.find('</tr>', header_start + 1)
    header_row = html_content[header_start:header_end]
    original_columns = header_row.count('<th>')
    print(f"  Original columns: {original_columns}")
    # Count rows
    row_count = html_content.count('<tr>') - 2  # Subtract header rows
    print(f"  Original rows: {row_count}")
except Exception as e:
    print(f"✗ Error: {e}")

# Test 2: Randomly add columns or rows
print("\nTest 2: Random changes")
try:
    # Load and modify the test file
    wb = openpyxl.load_workbook(test_file)
    ws = wb.active
    
    # Randomly decide to add rows or columns
    add_rows = random.choice([True, False])
    
    if add_rows:
        # Add random number of rows (2-5)
        num_rows = random.randint(2, 5)
        last_row = ws.max_row
        
        for i in range(num_rows):
            # Insert new row
            ws.insert_rows(last_row + i + 1)
            # Add data
            ws.cell(row=last_row + i + 1, column=1, value='测试项目')
            ws.cell(row=last_row + i + 1, column=2, value=f'测试名称{i+1}')
            ws.cell(row=last_row + i + 1, column=3, value='测试进展')
            ws.cell(row=last_row + i + 1, column=4, value='测试人员')
            ws.cell(row=last_row + i + 1, column=5, value='测试备注')
            ws.cell(row=last_row + i + 1, column=6, value='open')
        
        print(f"  Added {num_rows} random rows")
    else:
        # Add random number of columns (1-3)
        num_cols = random.randint(1, 3)
        last_col = ws.max_column
        
        for i in range(num_cols):
            # Insert new column
            ws.insert_cols(last_col + i + 1)
            # Add header
            new_header = f'新列{i+1}'
            ws.cell(row=2, column=last_col + i + 1, value=new_header)
            # Add data
            for row in range(3, ws.max_row + 1):
                ws.cell(row=row, column=last_col + i + 1, value=f'数据{row-2}')
        
        print(f"  Added {num_cols} random columns")
    
    # Save changes
    wb.save(test_file)
    
    # Test processing
    sender = WeeklyReportSender(excel_file_path=test_file)
    html_content = sender.read_excel_content()
    print("✓ Modified file processed successfully")
    
    # Check results
    if add_rows:
        # Check if new rows are in HTML
        if '测试项目' in html_content and '测试名称1' in html_content:
            print("✓ New rows are included in HTML")
        else:
            print("✗ New rows are not in HTML")
            
        # Count rows in modified file
        new_row_count = html_content.count('<tr>') - 2  # Subtract header rows
        print(f"  New row count: {new_row_count}")
        if new_row_count == row_count + num_rows:
            print("✓ Row count is correct")
        else:
            print("✗ Row count is incorrect")
    else:
        # Check if new columns are in HTML
        column_found = True
        for i in range(num_cols):
            if f'新列{i+1}' not in html_content:
                column_found = False
                break
        if column_found:
            print("✓ New columns are included in HTML")
        else:
            print("✗ New columns are not in HTML")
            
        # Count columns in modified file
        header_start = html_content.find('<tr>')
        header_end = html_content.find('</tr>', header_start + 1)
        header_row = html_content[header_start:header_end]
        new_column_count = header_row.count('<th>')
        print(f"  New column count: {new_column_count}")
        if new_column_count == original_columns + num_cols:
            print("✓ Column count is correct")
        else:
            print("✗ Column count is incorrect")
            
except Exception as e:
    print(f"✗ Error: {e}")

# Test 3: Test with both rows and columns added
print("\nTest 3: Add both rows and columns")
try:
    # Load and modify the test file
    wb = openpyxl.load_workbook(test_file)
    ws = wb.active
    
    # Add 2 columns
    num_cols = 2
    last_col = ws.max_column
    for i in range(num_cols):
        ws.insert_cols(last_col + i + 1)
        ws.cell(row=2, column=last_col + i + 1, value=f'额外列{i+1}')
        for row in range(3, ws.max_row + 1):
            ws.cell(row=row, column=last_col + i + 1, value=f'额外数据{row-2}')
    
    # Add 3 rows
    num_rows = 3
    last_row = ws.max_row
    for i in range(num_rows):
        ws.insert_rows(last_row + i + 1)
        ws.cell(row=last_row + i + 1, column=1, value='最终测试项目')
        ws.cell(row=last_row + i + 1, column=2, value=f'最终测试名称{i+1}')
        ws.cell(row=last_row + i + 1, column=3, value='最终测试进展')
        ws.cell(row=last_row + i + 1, column=4, value='最终测试人员')
        ws.cell(row=last_row + i + 1, column=5, value='最终测试备注')
        ws.cell(row=last_row + i + 1, column=6, value='open')
        # Add data to new columns
        for col in range(last_col + 1, last_col + num_cols + 1):
            ws.cell(row=last_row + i + 1, column=col, value=f'额外数据{last_row + i + 1}')
    
    wb.save(test_file)
    print("  Added 2 columns and 3 rows")
    
    # Test processing
    sender = WeeklyReportSender(excel_file_path=test_file)
    html_content = sender.read_excel_content()
    print("✓ File with both rows and columns added processed successfully")
    
    # Check results
    if '最终测试项目' in html_content and '额外列1' in html_content:
        print("✓ Both new rows and columns are included in HTML")
    else:
        print("✗ Some changes are not included in HTML")
        
except Exception as e:
    print(f"✗ Error: {e}")

# Clean up
if os.path.exists(test_file):
    os.remove(test_file)
    print(f"\nCleaned up test file: {test_file}")

print("\nAll tests completed!")