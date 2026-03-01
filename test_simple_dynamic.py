#!/usr/bin/env python3
"""
Simple test script to verify the system can handle dynamic Excel files.
"""

import os
import openpyxl
from weekly_report_sender import WeeklyReportSender

print("Testing dynamic Excel file handling...")

# Create a simple test Excel file
test_file = 'test_dynamic.xlsx'

# Load the original file and save as test file
wb = openpyxl.load_workbook('weekly_report.xlsx')
wb.save(test_file)
print(f"Created test file: {test_file}")

# Test 1: Original file
print("\nTest 1: Original file")
try:
    sender = WeeklyReportSender(excel_file_path=test_file)
    html_content = sender.read_excel_content()
    print("✓ Original file processed successfully")
    # Count columns in HTML
    table_start = html_content.find('<table')
    header_start = html_content.find('<thead>', table_start)
    header_end = html_content.find('</thead>', header_start)
    header_html = html_content[header_start:header_end]
    column_count = header_html.count('<th>') - 1  # Subtract the title row
    print(f"  Columns found: {column_count}")
except Exception as e:
    print(f"✗ Error processing original file: {e}")

# Test 2: Add a column
print("\nTest 2: Add a column")
try:
    wb = openpyxl.load_workbook(test_file)
    ws = wb.active
    # Add a new column after the last column
    last_col = ws.max_column
    ws.insert_cols(last_col + 1)
    # Add header
    ws.cell(row=2, column=last_col + 1, value='测试列')
    # Add some data
    for row in range(3, ws.max_row + 1):
        ws.cell(row=row, column=last_col + 1, value=f'测试数据{row-2}')
    wb.save(test_file)
    print("  Added new column '测试列'")
    
    # Test processing
    sender = WeeklyReportSender(excel_file_path=test_file)
    html_content = sender.read_excel_content()
    print("✓ File with added column processed successfully")
    # Check if new column is in HTML
    if '测试列' in html_content:
        print("✓ New column '测试列' is included in the HTML")
    else:
        print("✗ New column '测试列' is not included in the HTML")
except Exception as e:
    print(f"✗ Error processing file with added column: {e}")

# Test 3: Add rows
print("\nTest 3: Add rows")
try:
    wb = openpyxl.load_workbook(test_file)
    ws = wb.active
    # Add 2 new rows
    last_row = ws.max_row
    for i in range(2):
        ws.insert_rows(last_row + i + 1)
        # Add data
        ws.cell(row=last_row + i + 1, column=1, value='测试项目')
        ws.cell(row=last_row + i + 1, column=2, value=f'测试名称{i+1}')
        ws.cell(row=last_row + i + 1, column=3, value='测试进展')
        ws.cell(row=last_row + i + 1, column=4, value='测试人员')
        ws.cell(row=last_row + i + 1, column=5, value='测试备注')
        ws.cell(row=last_row + i + 1, column=6, value='open')
        ws.cell(row=last_row + i + 1, column=7, value=f'测试数据{last_row + i + 1}')
    wb.save(test_file)
    print("  Added 2 new rows")
    
    # Test processing
    sender = WeeklyReportSender(excel_file_path=test_file)
    html_content = sender.read_excel_content()
    print("✓ File with added rows processed successfully")
    # Check if new rows are in HTML
    if '测试项目' in html_content and '测试名称1' in html_content:
        print("✓ New rows are included in the HTML")
    else:
        print("✗ New rows are not included in the HTML")
except Exception as e:
    print(f"✗ Error processing file with added rows: {e}")

# Clean up
if os.path.exists(test_file):
    os.remove(test_file)
    print(f"\nCleaned up test file: {test_file}")

print("\nTest completed!")