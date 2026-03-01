#!/usr/bin/env python3
"""
Simple verification test for dynamic Excel handling.
"""

import os
import openpyxl
from weekly_report_sender import WeeklyReportSender

print("=== Testing Dynamic Excel Handling ===")
print()

# Test 1: Original file
print("Test 1: Processing original Excel file")
try:
    sender = WeeklyReportSender()
    html_content = sender.read_excel_content()
    print("✓ Original file processed successfully")
    
    # Check for existing columns
    if '项目' in html_content and '名称' in html_content and '进展' in html_content:
        print("✓ All original columns found")
    else:
        print("✗ Some original columns missing")
        
except Exception as e:
    print(f"✗ Error: {e}")

print()

# Test 2: Create a modified Excel file with extra column
print("Test 2: Processing Excel file with extra column")
try:
    # Create a copy of the original file
    wb = openpyxl.load_workbook('weekly_report.xlsx')
    ws = wb.active
    
    # Add a new column
    last_col = ws.max_column
    ws.insert_cols(last_col + 1)
    ws.cell(row=2, column=last_col + 1, value='测试列')
    
    # Add data to the new column
    for row in range(3, ws.max_row + 1):
        ws.cell(row=row, column=last_col + 1, value=f'测试数据{row-2}')
    
    # Save the modified file
    test_file = 'test_with_extra_column.xlsx'
    wb.save(test_file)
    print("  Added extra column '测试列'")
    
    # Process the modified file
    sender = WeeklyReportSender(excel_file_path=test_file)
    html_content = sender.read_excel_content()
    print("✓ Modified file processed successfully")
    
    # Check if the new column is included
    if '测试列' in html_content:
        print("✓ New column '测试列' is included in the output")
    else:
        print("✗ New column '测试列' is not included in the output")
        
    # Clean up
    if os.path.exists(test_file):
        os.remove(test_file)
        print("  Cleaned up test file")
        
except Exception as e:
    print(f"✗ Error: {e}")

print()
print("=== Test Completed ===")