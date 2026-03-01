#!/usr/bin/env python3
"""
Automated Weekly Report Sender
Reads Excel file content and sends it via email to configured recipients.
"""

import os
import sys
import logging
import smtplib
import pandas as pd
import openpyxl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('weekly_report.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

try:
    import pandas as pd
    import openpyxl
except ImportError:
    logger.error("Required libraries missing. Please install: pip install pandas openpyxl")
    sys.exit(1)

try:
    from config import EMAIL_CONFIG, RECIPIENTS_CONFIG, FILE_CONFIG
except ImportError:
    logger.error("config.py file not found. Please create the configuration file first.")
    sys.exit(1)


class WeeklyReportSender:
    """Handles reading Excel files and sending weekly reports via email."""
    
    def __init__(self):
        self.sender_email = EMAIL_CONFIG['sender_email']
        self.sender_password = EMAIL_CONFIG['sender_password']
        self.smtp_server = EMAIL_CONFIG['smtp_server']
        self.smtp_port = EMAIL_CONFIG['smtp_port']
        self.to_emails = RECIPIENTS_CONFIG['to_emails']
        self.cc_emails = RECIPIENTS_CONFIG['cc_emails']
        self.excel_file_path = FILE_CONFIG['excel_file_path']
        self.subject = FILE_CONFIG['subject']
        self.body_template = FILE_CONFIG['body_template']
    
    def validate_config(self):
        """Validate that all required configuration values are set."""
        if not self.sender_email or self.sender_email == 'your_email@example.com':
            raise ValueError("Please configure sender_email in config.py")
        if not self.sender_password or self.sender_password == 'your_app_password':
            raise ValueError("Please configure sender_password in config.py")
        if not self.to_emails or self.to_emails == ['recipient1@example.com', 'recipient2@example.com']:
            raise ValueError("Please configure to_emails in config.py")
        if not os.path.exists(self.excel_file_path):
            # Check for alternative filename without typo
            alt_filename = self.excel_file_path.replace('weekyly', 'weekly')
            if os.path.exists(alt_filename):
                self.excel_file_path = alt_filename
                logger.info(f"Using alternative filename: {alt_filename}")
            else:
                raise FileNotFoundError(f"Excel file not found: {self.excel_file_path}")
    
    def read_excel_with_merged_cells(self, file_path):
        """
        Read Excel file and handle merged cells properly.
        This function fills in the missing values that result from merged cells.
        """
        # Load the workbook
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # Get merged cells
        merged_cells = ws.merged_cells.ranges
        
        # Create a dictionary to hold merged cell values
        merged_values = {}
        for merged_range in merged_cells:
            # Get the top-left cell value
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_cell = ws.cell(row=min_row, column=min_col)
            value = top_left_cell.value
            
            # Store value for all cells in the merged range
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    merged_values[(row, col)] = value
        
        # Get the data
        data_rows = []
        headers = None
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 1:
                # Skip the title row "周报"
                continue
            elif row_idx == 2:
                # This is the header row
                headers = [cell if cell is not None else f'Column_{i}' for i, cell in enumerate(row)]
            else:
                # Data rows with merged cell handling
                processed_row = []
                for col_idx, cell_value in enumerate(row, 1):
                    # Check if this cell is part of a merged range
                    if (row_idx, col_idx) in merged_values:
                        processed_row.append(merged_values[(row_idx, col_idx)])
                    else:
                        processed_row.append(cell_value)
                data_rows.append(processed_row)
        
        wb.close()
        
        if headers is None:
            raise ValueError("Could not find header row in Excel file")
        
        # Create DataFrame
        df = pd.DataFrame(data_rows, columns=headers)
        
        # Handle any remaining NaN values
        df = df.fillna('')
        
        # Handle merged cells by forward-filling the '项目' column
        # This will propagate category names down through empty cells
        if '项目' in df.columns:
            df['项目'] = df['项目'].ffill()  # Use ffill() instead of fillna(method='ffill')
        
        return df
    
    def parse_excel_structure(self, df):
        """
        Parse the hierarchical Excel structure to extract meaningful content.
        After handling merged cells, the '项目' column should contain category names
        for all relevant rows.
        """
        parsed_content = []
        current_category = None
        
        # Group by category in the '项目' column
        for idx, row in df.iterrows():
            project_val = row['项目'] if pd.notna(row['项目']) else None
            name_val = row['名称'] if pd.notna(row['名称']) else None
            
            # Skip rows with no name (likely empty or category-only rows)
            if name_val is None:
                continue
            
            # Check if we have a new category
            if project_val is not None and project_val != current_category:
                current_category = project_val
                parsed_content.append(f"\n## {current_category}")
            
            # Extract other column values
            progress_val = row['进展'] if pd.notna(row['进展']) else ''
            handler_val = row['处理人'] if pd.notna(row['处理人']) else ''
            status_val = row['状态'] if pd.notna(row['状态']) else ''
            
            # Format the data row
            data_row = f"- **{name_val}**"
            if progress_val:
                data_row += f" | Progress: {progress_val}"
            if handler_val:
                data_row += f" | Handler: {handler_val}"
            if status_val:
                data_row += f" | Status: {status_val}"
            
            parsed_content.append(data_row)
        
        return "\n".join(parsed_content) if parsed_content else "No data found in Excel file."
    
    def read_excel_content(self):
        """Read and return the content of the Excel file as an HTML string with enhanced formatting, preserving cell styles and merged cells."""
        try:
            # Load the workbook with openpyxl to access formatting (data_only=False to preserve formatting)
            wb = openpyxl.load_workbook(self.excel_file_path, data_only=False)
            ws = wb.active
            
            # Get merged cell ranges
            merged_cells = list(ws.merged_cells.ranges)
            
            # Create a dictionary to hold merged cell information
            merged_info = {}
            for merged_range in merged_cells:
                min_col, min_row, max_col, max_row = merged_range.bounds
                merged_info[(min_row, min_col)] = (max_row, max_col)
            
            # Read all columns from the Excel file
            column_indices = []
            headers = []
            
            # Check row 2 for headers
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=2, column=col)
                header_value = cell.value
                if header_value:
                    column_indices.append(col)
                    headers.append(header_value)
            
            # If no headers found, use default headers
            if not headers:
                column_indices = list(range(1, min(6, ws.max_column + 1)))
                headers = ['项目', '名称', '进展', '处理人', '状态'][:len(column_indices)]
            
            # Build HTML table
            html = ['<table class="excel-table" border="1" cellspacing="0" cellpadding="4">']
            
            # Add title row (周报)
            title_cell = ws.cell(row=1, column=1)
            title_value = title_cell.value if title_cell.value is not None else '周报'
            html.append('  <thead>')
            html.append('    <tr>')
            html.append(f'      <th colspan="{len(headers)}" style="text-align: center; font-weight: bold; font-size: 14pt;">{title_value}</th>')
            html.append('    </tr>')
            
            # Add header row
            html.append('    <tr>')
            for header in headers:
                html.append(f'      <th>{header}</th>')
            html.append('    </tr>')
            html.append('  </thead>')
            
            # Add data rows
            html.append('  <tbody>')
            
            # Track merged cells that span multiple rows
            merged_rows = {}
            
            for row_idx in range(3, ws.max_row + 1):  # Start from row 3 (data rows)
                html.append('    <tr>')
                col_index = 0
                
                while col_index < len(column_indices):
                    col_idx = column_indices[col_index]
                    
                    # Check if this cell is part of a merged range that started in a previous row
                    skip_cell = False
                    for (start_row, start_col), (end_row, end_col) in merged_info.items():
                        if start_row < row_idx <= end_row and start_col <= col_idx <= end_col:
                            # This cell is part of a merged range from a previous row
                            skip_cell = True
                            break
                    
                    if skip_cell:
                        # Move to the next column
                        col_index += 1
                        continue
                    
                    # Check if this cell is the top-left of a new merged range
                    if (row_idx, col_idx) in merged_info:
                        max_row, max_col = merged_info[(row_idx, col_idx)]
                        rowspan = max_row - row_idx + 1
                        # Calculate colspan within our desired columns
                        colspan = 1
                        # Only count columns that are in our column_indices list
                        for i in range(col_idx + 1, max_col + 1):
                            if i in column_indices:
                                colspan += 1
                        # Ensure colspan doesn't exceed the number of columns we're displaying
                        colspan = min(colspan, len(column_indices) - col_index)
                    else:
                        rowspan = 1
                        colspan = 1
                    
                    # Get cell value
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value if cell.value is not None else ''
                    
                    # Get cell styling - preserve original formatting from Excel
                    style_attrs = []
                    
                    # Background color
                    try:
                        # Check if cell has fill
                        if cell.fill and cell.fill.start_color:
                            # Get RGB value
                            fill_color = cell.fill.start_color
                            if fill_color and hasattr(fill_color, 'rgb'):
                                # Check if rgb is a string
                                if isinstance(fill_color.rgb, str):
                                    rgb_str = fill_color.rgb
                                    # Extract RGB part (remove alpha channel if present)
                                    if len(rgb_str) == 8:
                                        hex_color = '#' + rgb_str[2:]
                                    else:
                                        hex_color = '#' + rgb_str
                                    # Only set background color if it's not black (for readability)
                                    if hex_color != '#000000':
                                        style_attrs.append(f'background-color: {hex_color}')
                    except Exception as e:
                        logger.debug(f"Error processing background color: {e}")
                    
                    # Font color
                    try:
                        # Check if cell has font color
                        if cell.font.color:
                            font_color = cell.font.color
                            if font_color:
                                # Handle theme-based colors
                                if hasattr(font_color, 'theme') and font_color.theme is not None:
                                    # For theme-based colors, use default black (theme 1 is usually black)
                                    if font_color.theme == 1:
                                        style_attrs.append('color: #000000')
                                # Handle direct RGB colors
                                elif hasattr(font_color, 'rgb') and font_color.rgb:
                                    rgb_str = str(font_color.rgb)
                                    # Extract RGB part (remove alpha channel if present)
                                    if len(rgb_str) == 8:
                                        hex_color = '#' + rgb_str[2:]
                                    else:
                                        hex_color = '#' + rgb_str
                                    style_attrs.append(f'color: {hex_color}')
                    except Exception as e:
                        logger.debug(f"Error processing font color: {e}")
                    
                    # Font size
                    if cell.font.size:
                        try:
                            style_attrs.append(f'font-size: {cell.font.size}pt')
                        except Exception as e:
                            logger.debug(f"Error processing font size: {e}")
                    
                    # Font weight
                    if cell.font.bold:
                        style_attrs.append('font-weight: bold')
                    
                    # Font style
                    if cell.font.italic:
                        style_attrs.append('font-style: italic')
                    
                    # Alignment
                    if cell.alignment:
                        if cell.alignment.horizontal:
                            style_attrs.append(f'text-align: {cell.alignment.horizontal}')
                        if cell.alignment.vertical:
                            style_attrs.append(f'vertical-align: {cell.alignment.vertical}')
                    
                    # Build style attribute
                    style_str = '; '.join(style_attrs)
                    style_html = f' style="{style_str}"'
                    
                    # Build cell HTML
                    if rowspan > 1 or colspan > 1:
                        html.append(f'      <td{style_html} rowspan="{rowspan}" colspan="{colspan}">{value}</td>')
                    else:
                        html.append(f'      <td{style_html}>{value}</td>')
                    
                    # Move to the next column after the merged range
                    col_index += colspan
                html.append('    </tr>')
            html.append('  </tbody>')
            html.append('</table>')
            
            # Add CSS styling
            style = """
            <style>
                .excel-table {
                    border-collapse: collapse;
                    width: 100%;
                    font-family: Arial, sans-serif;
                    font-size: 12px;
                    margin: 0 auto;
                }
                .excel-table th, .excel-table td {
                    border: 1px solid #ddd;
                    padding: 8px;
                    text-align: left;
                }
                .excel-table th {
                    background-color: #f2f2f2;
                    font-weight: bold;
                    text-align: center;
                }
                .excel-table tr:nth-child(even) {
                    background-color: #f9f9f9;
                }
                .excel-table tr:hover {
                    background-color: #f5f5f5;
                }
            </style>
            """
            
            html_content = style + '\n'.join(html)
            
            # Log summary information
            logger.info(f"Successfully processed Excel file: {ws.max_row - 2} rows processed")
            logger.info(f"Found {len(merged_cells)} merged cell ranges")
            logger.info(f"Using {len(column_indices)} columns: {headers}")
            
            return html_content
            
        except Exception as e:
            logger.error(f"Error reading Excel file: {str(e)}")
            raise
    
    def create_email_message(self, excel_content):
        """Create the email message with formatted Excel content in the body."""
        msg = MIMEMultipart('alternative')
        msg['From'] = self.sender_email
        msg['To'] = ', '.join(self.to_emails)
        msg['Cc'] = ', '.join(self.cc_emails) if self.cc_emails else ''
        msg['Subject'] = self.subject
        
        # Convert plain text template to HTML
        html_template = self.body_template.replace('\n', '<br>')
        
        # Email body with HTML content
        html_body = f"""
        <html>
          <head></head>
          <body>
            <p>{html_template}</p>
            <hr>
            <h2>Weekly Report Content:</h2>
            {excel_content}
          </body>
        </html>
        """
        
        # Attach HTML part
        msg.attach(MIMEText(html_body, 'html'))
        
        return msg
    
    def send_email(self, msg):
        """Send the email using SMTP."""
        try:
            # Create SMTP session
            if self.smtp_port == 465:
                # Use SSL for port 465
                server = smtplib.SMTP_SSL(self.smtp_server, self.smtp_port)
            else:
                # Use TLS for other ports (25, 587)
                server = smtplib.SMTP(self.smtp_server, self.smtp_port)
                server.starttls()  # Enable TLS encryption
            
            server.login(self.sender_email, self.sender_password)
            
            # Get all recipients (to + cc)
            all_recipients = self.to_emails + (self.cc_emails if self.cc_emails else [])
            
            # Send email
            text = msg.as_string()
            server.sendmail(self.sender_email, all_recipients, text)
            server.quit()
            
            logger.info(f"Email sent successfully to {len(all_recipients)} recipients")
            logger.info(f"To: {', '.join(self.to_emails)}")
            if self.cc_emails:
                logger.info(f"CC: {', '.join(self.cc_emails)}")
                
        except smtplib.SMTPAuthenticationError:
            logger.error("SMTP authentication failed. Please check your email and password.")
            raise
        except smtplib.SMTPException as e:
            logger.error(f"SMTP error occurred: {str(e)}")
            raise
        except Exception as e:
            logger.error(f"Unexpected error while sending email: {str(e)}")
            raise
    
    def run(self):
        """Main execution method."""
        try:
            logger.info("Starting weekly report automation...")
            
            # Validate configuration
            self.validate_config()
            logger.info("Configuration validated successfully")
            
            # Read and parse Excel content with merged cell support
            excel_content = self.read_excel_content()
            
            # Create email message
            msg = self.create_email_message(excel_content)
            
            # Send email
            self.send_email(msg)
            
            logger.info("Weekly report automation completed successfully!")
            
        except Exception as e:
            logger.error(f"Weekly report automation failed: {str(e)}")
            raise


def main():
    """Main function to run the weekly report sender."""
    try:
        sender = WeeklyReportSender()
        sender.run()
    except KeyboardInterrupt:
        logger.info("Process interrupted by user")
        sys.exit(0)
    except Exception as e:
        logger.error(f"Application failed: {str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    main()